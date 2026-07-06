from flask import Flask, request, jsonify, render_template, send_file
import subprocess
import time
import re
import json
import asyncio
import websockets
import threading
import os
import io
import socket
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter

app = Flask(__name__, template_folder='.', static_folder='.', static_url_path='')

UPLOAD_FOLDER = os.path.dirname(os.path.abspath(__file__))
frame_history = []
battery_history = []
system_history = []
tracking_active = False
active_package_name = None
loop = None

latest_broadcast_event = "None"

def run_command(command):
    result = subprocess.run(command, shell=True, capture_output=True, text=True)
    if result.returncode != 0:
        return False, result.stderr
    return True, result.stdout

async def stream_performance_data(ws_url):
    global frame_history, tracking_active, latest_broadcast_event
    try:
        async with websockets.connect(ws_url) as websocket:
            await websocket.send(json.dumps({
                "jsonrpc": "2.0",
                "method": "streamListen",
                "params": {"streamId": "Extension"},
                "id": 1
            }))
            
            while tracking_active:
                try:
                    response = await asyncio.wait_for(websocket.recv(), timeout=1.0)
                    data = json.loads(response)
                    
                    if "params" in data and data["params"]["event"]["extensionKind"] == "Flutter.Frame":
                        frame_data = data["params"]["event"]["extensionData"]
                        raw_ui = frame_data.get("build", frame_data.get("uiDuration", 0))
                        raw_raster = frame_data.get("raster", frame_data.get("rasterDuration", 0))
                        
                        frame_history.append({
                            "frame_id": frame_data.get("number", "N/A"),
                            "ui_time_ms": round(raw_ui / 1000.0, 2),
                            "raster_time_ms": round(raw_raster / 1000.0, 2),
                            "total_time_ms": round((raw_ui + raw_raster) / 1000.0, 2),
                            "jank": ((raw_ui + raw_raster) / 1000.0) > 16.6,
                            "triggered_event": latest_broadcast_event,
                            "ts": time.time()
                        })
                        
                        latest_broadcast_event = "None"
                except asyncio.TimeoutError:
                    continue
    except Exception as e:
        print(f"WebSocket Error: {e}")

def listen_for_android_broadcasts():
    global latest_broadcast_event, tracking_active
    logcat_cmd = "adb logcat -v raw *:S DreadTelemetry:V"
    process = subprocess.Popen(logcat_cmd, shell=True, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True)
    
    while tracking_active:
        line = process.stdout.readline()
        if not line: break
        if "Event:" in line:
            try:
                latest_broadcast_event = line.split("Event:")[1].strip()
                print(f"[HEXA OS INTENT INTERCEPTED]: {latest_broadcast_event}")
            except Exception: pass
                
    process.terminate()

def poll_battery_hardware():
    global battery_history, tracking_active
    start_time = time.time()
    
    while tracking_active:
        success, output = run_command("adb shell dumpsys battery")
        if success:
            level_match = re.search(r'level:\s+(\d+)', output)
            temp_match = re.search(r'temperature:\s+(\d+)', output)
            
            level = int(level_match.group(1)) if level_match else 0
            temp_raw = int(temp_match.group(1)) if temp_match else 0
            temperature_c = round(temp_raw / 10.0, 1)
            
            battery_history.append({
                "time_sec": int(time.time() - start_time),
                "battery_level": level,
                "temperature_c": temperature_c
            })
        time.sleep(1.0)

def poll_system_stats():
    global system_history, tracking_active, active_package_name
    start_time = time.time()
    pkg = active_package_name

    target_pid = None
    prev_ticks = None
    prev_uptime = None
    num_cores = 1

    ok, nproc_out = run_command("adb shell nproc")
    if ok and nproc_out.strip().isdigit():
        num_cores = int(nproc_out.strip())

    while tracking_active:
        entry = {"time_sec": int(time.time() - start_time), "memory_mb": 0, "cpu_percent": 0}

        success, output = run_command(f"adb shell dumpsys meminfo {pkg} -s")
        if success:
            total_pss = re.search(r'TOTAL\s+(\d+)', output)
            if total_pss:
                entry["memory_mb"] = round(int(total_pss.group(1)) / 1024.0, 1)

        if target_pid is None:
            ok, ps_out = run_command(f"adb shell ps -A")
            if ok:
                for line in ps_out.splitlines():
                    if pkg in line:
                        parts = line.split()
                        if len(parts) >= 2 and parts[1].isdigit():
                            target_pid = parts[1]
                            break

        if target_pid:
            ok1, stat_out = run_command(f"adb shell cat /proc/{target_pid}/stat")
            ok2, up_out = run_command("adb shell cat /proc/uptime")
            if ok1 and ok2:
                fields = stat_out.split()
                if len(fields) >= 15:
                    cur_ticks = int(fields[13]) + int(fields[14])
                    cur_uptime = float(up_out.split()[0])

                    if prev_ticks is not None and prev_uptime is not None:
                        dt = cur_uptime - prev_uptime
                        dticks = cur_ticks - prev_ticks
                        if dt > 0:
                            entry["cpu_percent"] = round((dticks / 100.0) / dt * 100 / num_cores, 1)

                    prev_ticks = cur_ticks
                    prev_uptime = cur_uptime

        system_history.append(entry)
        time.sleep(2.0)

def run_async_loop(ws_url):
    global loop
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    loop.run_until_complete(stream_performance_data(ws_url))

@app.route('/')
def home():
    index_path = os.path.join(UPLOAD_FOLDER, 'index.html')
    with open(index_path, 'r', encoding='utf-8') as f:
        return f.read()

@app.route('/api/start', methods=['POST'])
def start_test():
    global frame_history, battery_history, system_history, tracking_active, latest_broadcast_event, active_package_name
    package_name = request.form.get('package_name')
    apk_file = request.files.get('apk')
    if not package_name or not apk_file: return jsonify({"success": False, "error": "Missing config properties."}), 400

    frame_history, battery_history, system_history, latest_broadcast_event, tracking_active = [], [], [], "None", True
    active_package_name = package_name
    apk_path = os.path.join(UPLOAD_FOLDER, "temp-target-profile.apk")
    apk_file.save(apk_path)

    success, devices = run_command("adb devices")
    if not success or len([l for l in devices.strip().split('\n') if l]) <= 1: return jsonify({"success": False, "error": "No USB devices detected."})

    success, err = run_command(f"adb install -r {apk_path}")
    if not success: return jsonify({"success": False, "error": f"Install failed: {err}"})
    
    run_command("adb logcat -c")
    success, err = run_command(f"adb shell monkey -p {package_name} -c android.intent.category.LAUNCHER 1")
    if not success: return jsonify({"success": False, "error": f"Launch fault: {err}"})

    logcat_process = subprocess.Popen("adb logcat", shell=True, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True)
    port, auth_token = None, None
    timeout = time.time() + 20
    
    while time.time() < timeout:
        line = logcat_process.stdout.readline()
        if "The Dart VM service is listening on" in line or "Observatory listening on" in line:
            match = re.search(r'http://127\.0\.0\.1:(\d+)/([^/\s]*)/', line)
            if match:
                port, auth_token = match.group(1), match.group(2)
                break
    logcat_process.terminate()

    if not port: return jsonify({"success": False, "error": "Could not extract Dart VM metrics configuration pipeline."})

    run_command(f"adb forward tcp:{port} tcp:{port}")
    ws_url = f"ws://127.0.0.1:{port}/{auth_token}/ws"
    
    threading.Thread(target=run_async_loop, args=(ws_url,), daemon=True).start()
    threading.Thread(target=listen_for_android_broadcasts, daemon=True).start()
    threading.Thread(target=poll_battery_hardware, daemon=True).start()
    threading.Thread(target=poll_system_stats, daemon=True).start()
    
    return jsonify({"success": True, "message": "All bulletproof layers connected live!"})

@app.route('/api/live', methods=['GET'])
def get_live_data():
    global frame_history, battery_history, system_history, tracking_active
    frame_idx = int(request.args.get('frame_start', 0))
    battery_idx = int(request.args.get('battery_start', 0))
    system_idx = int(request.args.get('system_start', 0))

    recent_count = min(120, len(frame_history))
    if recent_count >= 2:
        recent = frame_history[-recent_count:]
        elapsed = recent[-1]["ts"] - recent[0]["ts"]
        fps = round((recent_count - 1) / elapsed, 1) if elapsed > 0 else 0
    else:
        fps = 0

    return jsonify({
        "active": tracking_active,
        "new_frames": frame_history[frame_idx:],
        "new_battery": battery_history[battery_idx:],
        "new_system": system_history[system_idx:],
        "fps": fps
    })

@app.route('/api/stop', methods=['POST'])
def stop_test():
    global tracking_active, frame_history, battery_history, system_history
    tracking_active = False
    
    if not frame_history: return jsonify({"success": False, "error": "No data captured."})

    total_frames = len(frame_history)
    janky_frames = sum(1 for f in frame_history if f["jank"])
    jank_percentage = (janky_frames / total_frames) * 100 if total_frames > 0 else 0

    avg_mem = round(sum(s["memory_mb"] for s in system_history) / len(system_history), 1) if system_history else 0
    peak_mem = max((s["memory_mb"] for s in system_history), default=0)
    avg_cpu = round(sum(s["cpu_percent"] for s in system_history) / len(system_history), 1) if system_history else 0
    peak_cpu = max((s["cpu_percent"] for s in system_history), default=0)

    report = {
        "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
        "summary": {
            "total_frames_captured": total_frames,
            "total_janky_frames": janky_frames,
            "jank_percentage": round(jank_percentage, 2),
            "average_ui_time_ms": round(sum(f["ui_time_ms"] for f in frame_history) / total_frames, 2),
            "average_raster_time_ms": round(sum(f["raster_time_ms"] for f in frame_history) / total_frames, 2),
            "health_score": round(100 - jank_percentage, 1),
            "avg_memory_mb": avg_mem,
            "peak_memory_mb": peak_mem,
            "avg_cpu_percent": avg_cpu,
            "peak_cpu_percent": peak_cpu
        },
        "frames": frame_history,
        "battery": battery_history,
        "system": system_history
    }
    return jsonify({"success": True, "report": report})

def calculate_action_spans():
    spans = []
    last_evt = None
    last_idx = None
    
    for idx, f in enumerate(frame_history):
        evt = f.get("triggered_event", "None")
        if evt and evt != "None":
            if last_evt:
                span_frames = frame_history[last_idx:idx+1]
                total_in_span = len(span_frames)
                max_ui = max((sf["ui_time_ms"] for sf in span_frames), default=0)
                max_raster = max((sf["raster_time_ms"] for sf in span_frames), default=0)
                
                janky_in_span = sum(1 for sf in span_frames if sf["jank"])
                jank_ratio = round((janky_in_span / total_in_span) * 100, 1) if total_in_span > 0 else 0
                span_health = round(100 - jank_ratio, 1)

                status = "✅ OK"
                if max_ui > 16.6 or max_raster > 16.6: status = "⚠️ JANK IN SPAN"
                if max_ui > 32.0 or max_raster > 32.0: status = "❌ SEVERE JANK"

                spans.append({
                    "start_event": last_evt,
                    "end_event": evt,
                    "start_frame": frame_history[last_idx]["frame_id"],
                    "end_frame": f["frame_id"],
                    "duration_frames": total_in_span,
                    "janky_frames": janky_in_span,
                    "jank_ratio": jank_ratio,
                    "span_health": span_health,
                    "peak_ui": max_ui,
                    "peak_raster": max_raster,
                    "status": status
                })
            last_evt = evt
            last_idx = idx
    return spans

@app.route('/api/export/excel', methods=['GET'])
def export_excel():
    global frame_history, battery_history, system_history
    if not frame_history: return jsonify({"success": False, "error": "No session history data"}), 400

    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_summary.title = "Dashboard Summary"
    ws_summary.views.sheetView[0].showGridLines = False
    
    ws_spans = wb.create_sheet(title="AI Breakpoint Spans")
    ws_events = wb.create_sheet(title="Event RCA Diagnostics")
    ws_data = wb.create_sheet(title="Frame Logs")
    ws_batt = wb.create_sheet(title="Battery Logs")
    ws_sys = wb.create_sheet(title="System Stats")

    PRIMARY_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_alert = Font(name="Segoe UI", size=11, bold=True, color="DC2626")
    font_link = Font(name="Segoe UI", size=11, underline="single", color="3B82F6")

    # 1. Raw Frames Data
    headers = ["Frame ID", "UI Time (ms)", "Raster Time (ms)", "Total Time (ms)", "Triggered Game Event"]
    for col_num, h in enumerate(headers, 1):
        c = ws_data.cell(row=1, column=col_num, value=h)
        c.fill = PRIMARY_FILL; c.font = font_header

    for idx, f in enumerate(frame_history, 2):
        ws_data.cell(row=idx, column=1, value=f"F-{f['frame_id']}")
        ws_data.cell(row=idx, column=2, value=f["ui_time_ms"])
        ws_data.cell(row=idx, column=3, value=f["raster_time_ms"])
        ws_data.cell(row=idx, column=4, value=f["total_time_ms"])
        ws_data.cell(row=idx, column=5, value=f.get("triggered_event", "None"))

    # 2. Breakpoint Spans
    span_headers = [
        "Start Event", "End Event", "Start Frame", "End Frame",
        "Total Frames", "Janky Frames", "Jank Ratio (%)", "Span Health",
        "Peak UI (ms)", "Peak Raster (ms)",
        "Span Status"
    ]
    for col_num, h in enumerate(span_headers, 1):
        c = ws_spans.cell(row=1, column=col_num, value=h)
        c.fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid"); c.font = font_header

    font_good = Font(name="Segoe UI", size=11, bold=True, color="10B981")

    spans = calculate_action_spans()
    if not spans:
        ws_spans.merge_cells('A2:K2')
        cell = ws_spans.cell(row=2, column=1, value="⚠️ No contiguous DreadTelemetry events were intercepted. AI Breakpoint Spans require at least two events.")
        cell.font = Font(italic=True, color="6B7280"); cell.alignment = Alignment(horizontal="center")
    else:
        for idx, sp in enumerate(spans, 2):
            ws_spans.cell(row=idx, column=1, value=sp["start_event"])
            ws_spans.cell(row=idx, column=2, value=sp["end_event"])
            ws_spans.cell(row=idx, column=3, value=f"F-{sp['start_frame']}")
            ws_spans.cell(row=idx, column=4, value=f"F-{sp['end_frame']}")
            ws_spans.cell(row=idx, column=5, value=sp["duration_frames"])
            ws_spans.cell(row=idx, column=6, value=sp["janky_frames"])

            c_ratio = ws_spans.cell(row=idx, column=7, value=f"{sp['jank_ratio']}%")
            if sp["jank_ratio"] > 10: c_ratio.font = font_alert

            c_health = ws_spans.cell(row=idx, column=8, value=f"{sp['span_health']}/100")
            if sp["span_health"] >= 90: c_health.font = font_good
            elif sp["span_health"] < 80: c_health.font = font_alert

            ws_spans.cell(row=idx, column=9, value=sp["peak_ui"])
            ws_spans.cell(row=idx, column=10, value=sp["peak_raster"])

            c_stat = ws_spans.cell(row=idx, column=11, value=sp["status"])
            if "JANK" in sp["status"]: c_stat.font = font_alert

    # 3. RCA Events
    event_headers = ["Frame ID", "Game Event Triggered", "UI Time", "Raster Time", "Threshold Status", "Interact"]
    for col_num, h in enumerate(event_headers, 1):
        c = ws_events.cell(row=1, column=col_num, value=h)
        c.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid"); c.font = font_header

    event_row_counter = 2
    events_found = False
    for idx, f in enumerate(frame_history, 2):
        event_str = f.get("triggered_event", "None")
        if event_str and event_str != "None":
            events_found = True
            ui_t, ras_t = f["ui_time_ms"], f["raster_time_ms"]
            if ui_t > 16.6 and ras_t > 16.6: status = "❌ BOTH CRASHED BUDGET"
            elif ui_t > 16.6: status = "⚠️ UI THREAD BOTTLENECK"
            elif ras_t > 16.6: status = "⚠️ GPU THREAD BOTTLENECK"
            else: status = "✅ PERFECT"

            ws_events.cell(row=event_row_counter, column=1, value=f"F-{f['frame_id']}")
            ws_events.cell(row=event_row_counter, column=2, value=event_str)
            ws_events.cell(row=event_row_counter, column=3, value=ui_t)
            ws_events.cell(row=event_row_counter, column=4, value=ras_t)
            c_status = ws_events.cell(row=event_row_counter, column=5, value=status)
            if "BOTTLENECK" in status or "CRASHED" in status: c_status.font = font_alert

            c_link = ws_events.cell(row=event_row_counter, column=6, value="Jump to Frame →")
            c_link.hyperlink = f"#'Frame Logs'!A{idx}"; c_link.font = font_link
            event_row_counter += 1

    if not events_found:
        ws_events.merge_cells('A2:F2')
        cell = ws_events.cell(row=2, column=1, value="⚠️ No DreadTelemetry events were intercepted. Ensure your MethodChannel is firing correctly from your game.")
        cell.font = Font(italic=True, color="6B7280"); cell.alignment = Alignment(horizontal="center")

    # 4. Battery Logs
    b_headers = ["Timeline (Seconds)", "Charge Capacity (%)", "Temperature (C)"]
    for col_num, h in enumerate(b_headers, 1):
        c = ws_batt.cell(row=1, column=col_num, value=h)
        c.fill = PRIMARY_FILL; c.font = font_header

    for idx, b in enumerate(battery_history, 2):
        ws_batt.cell(row=idx, column=1, value=f"{b['time_sec']}s")
        ws_batt.cell(row=idx, column=2, value=b["battery_level"])
        ws_batt.cell(row=idx, column=3, value=b["temperature_c"])

    # 5. System Stats Logs
    sys_headers = ["Timeline (Seconds)", "Memory PSS (MB)", "CPU (%)"]
    for col_num, h in enumerate(sys_headers, 1):
        c = ws_sys.cell(row=1, column=col_num, value=h)
        c.fill = PRIMARY_FILL; c.font = font_header

    for idx, s in enumerate(system_history, 2):
        ws_sys.cell(row=idx, column=1, value=f"{s['time_sec']}s")
        ws_sys.cell(row=idx, column=2, value=s["memory_mb"])
        ws_sys.cell(row=idx, column=3, value=s["cpu_percent"])

    ws_sys.column_dimensions['A'].width = 20
    ws_sys.column_dimensions['B'].width = 18
    ws_sys.column_dimensions['C'].width = 12

    # 6. Dashboard Summary
    ws_summary.cell(row=2, column=2, value="Enterprise Performance RCA Report").font = Font(name="Segoe UI", size=18, bold=True, color="10B981")
    
    total_frames = len(frame_history)
    janky_frames = sum(1 for f in frame_history if f["jank"])
    avg_ui = round(sum(f["ui_time_ms"] for f in frame_history) / total_frames, 2) if total_frames > 0 else 0
    avg_raster = round(sum(f["raster_time_ms"] for f in frame_history) / total_frames, 2) if total_frames > 0 else 0
    jank_pct = round((janky_frames / total_frames) * 100, 2) if total_frames > 0 else 0
    health = round(100 - jank_pct, 1)

    avg_mem = round(sum(s["memory_mb"] for s in system_history) / len(system_history), 1) if system_history else 0
    peak_mem = max((s["memory_mb"] for s in system_history), default=0)
    avg_cpu = round(sum(s["cpu_percent"] for s in system_history) / len(system_history), 1) if system_history else 0
    peak_cpu = max((s["cpu_percent"] for s in system_history), default=0)

    kpi_labels = ["SYSTEM HEALTH SCORE", "OVERALL JANK RATIO", "AVG UI THREAD", "AVG RASTER THREAD"]
    kpi_vals = [f"{health}/100", f"{jank_pct}%", f"{avg_ui}ms", f"{avg_raster}ms"]
    
    for i, (lbl, val) in enumerate(zip(kpi_labels, kpi_vals)):
        col = 2 + (i * 2)
        c_lbl = ws_summary.cell(row=4, column=col, value=lbl)
        c_lbl.font = Font(name="Segoe UI", size=10, bold=True, color="6B7280"); c_lbl.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        c_val = ws_summary.cell(row=5, column=col, value=val)
        c_val.font = Font(name="Segoe UI", size=16, bold=True, color="1F2937"); c_val.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")

    kpi2_labels = ["AVG MEMORY (MB)", "PEAK MEMORY (MB)", "AVG CPU (%)", "PEAK CPU (%)"]
    kpi2_vals = [f"{avg_mem}", f"{peak_mem}", f"{avg_cpu}%", f"{peak_cpu}%"]
    for i, (lbl, val) in enumerate(zip(kpi2_labels, kpi2_vals)):
        col = 2 + (i * 2)
        c_lbl = ws_summary.cell(row=6, column=col, value=lbl)
        c_lbl.font = Font(name="Segoe UI", size=10, bold=True, color="6B7280"); c_lbl.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        c_val = ws_summary.cell(row=7, column=col, value=val)
        c_val.font = Font(name="Segoe UI", size=16, bold=True, color="1F2937"); c_val.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")

    dynamic_chart_width = max(25, int(total_frames * 0.10))

    # Frame Timeline Chart
    chart1 = LineChart()
    chart1.title = "Playthrough Frame Timeline Analysis"
    chart1.width = dynamic_chart_width
    chart1.height = 12
    data_ref = Reference(ws_data, min_col=2, min_row=1, max_col=3, max_row=total_frames+1)
    cats_ref = Reference(ws_data, min_col=1, min_row=2, max_row=total_frames+1)
    chart1.add_data(data_ref, titles_from_data=True)
    chart1.set_categories(cats_ref)
    chart1.series[0].graphicalProperties.line.solidFill = "3B82F6"
    chart1.series[1].graphicalProperties.line.solidFill = "A855F7"
    ws_summary.add_chart(chart1, "B8")

    # Battery Timeline Chart
    if battery_history:
        chart2 = LineChart()
        chart2.title = "Hardware Power & Thermals Timeline"
        chart2.width = dynamic_chart_width
        chart2.height = 10
        b_data_ref = Reference(ws_batt, min_col=2, min_row=1, max_col=3, max_row=len(battery_history)+1)
        b_cats_ref = Reference(ws_batt, min_col=1, min_row=2, max_row=len(battery_history)+1)
        chart2.add_data(b_data_ref, titles_from_data=True)
        chart2.set_categories(b_cats_ref)
        chart2.series[0].graphicalProperties.line.solidFill = "10B981"
        chart2.series[1].graphicalProperties.line.solidFill = "F97316"
        ws_summary.add_chart(chart2, "B31")

    # System Stats Chart (Memory + CPU)
    if system_history:
        chart3 = LineChart()
        chart3.title = "Memory & CPU Timeline"
        chart3.width = dynamic_chart_width
        chart3.height = 10
        s_data_ref = Reference(ws_sys, min_col=2, min_row=1, max_col=3, max_row=len(system_history)+1)
        s_cats_ref = Reference(ws_sys, min_col=1, min_row=2, max_row=len(system_history)+1)
        chart3.add_data(s_data_ref, titles_from_data=True)
        chart3.set_categories(s_cats_ref)
        chart3.series[0].graphicalProperties.line.solidFill = "EF4444"
        chart3.series[1].graphicalProperties.line.solidFill = "F59E0B"
        ws_summary.add_chart(chart3, "B54")

    # Formatting Column Widths
    for col in ["B", "D", "F", "H"]: ws_summary.column_dimensions[col].width = 22
    ws_data.column_dimensions['A'].width = 15
    ws_data.column_dimensions['E'].width = 35
    ws_events.column_dimensions['B'].width = 35
    ws_spans.column_dimensions['A'].width = 30
    ws_spans.column_dimensions['B'].width = 30
    ws_spans.column_dimensions['G'].width = 14
    ws_spans.column_dimensions['H'].width = 14
    ws_batt.column_dimensions['A'].width = 20

    excel_stream = io.BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)
    
    return send_file(excel_stream, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=f"HEXA_RCA_{time.strftime('%Y%m%d_%H%M%S')}.xlsx")


@app.route('/api/export/md', methods=['GET'])
def export_md():
    global frame_history, battery_history, system_history
    if not frame_history: return jsonify({"success": False, "error": "No session history data"}), 400

    total_frames = len(frame_history)
    janky_frames = sum(1 for f in frame_history if f["jank"])
    avg_ui = round(sum(f["ui_time_ms"] for f in frame_history) / total_frames, 2) if total_frames > 0 else 0
    avg_raster = round(sum(f["raster_time_ms"] for f in frame_history) / total_frames, 2) if total_frames > 0 else 0
    jank_pct = round((janky_frames / total_frames) * 100, 2) if total_frames > 0 else 0
    health = round(100 - jank_pct, 1)

    avg_mem = round(sum(s["memory_mb"] for s in system_history) / len(system_history), 1) if system_history else 0
    peak_mem = max((s["memory_mb"] for s in system_history), default=0)
    avg_cpu = round(sum(s["cpu_percent"] for s in system_history) / len(system_history), 1) if system_history else 0
    peak_cpu = max((s["cpu_percent"] for s in system_history), default=0)

    md = []
    md.append("# HEXA Performance Observability Lab - Executive AI Audit")
    md.append(f"**Exported On:** {time.strftime('%Y-%m-%d %H:%M:%S')}\n")

    md.append("## 1. High-Level Telemetry Summary")
    md.append(f"- **System Health Score:** {health}/100")
    md.append(f"- **Overall Jank Ratio:** {jank_pct}%")
    md.append(f"- **Total Frames Captured:** {total_frames}")
    md.append(f"- **Avg UI Thread:** {avg_ui} ms")
    md.append(f"- **Avg Raster Thread:** {avg_raster} ms")
    md.append(f"- **Avg Memory (PSS):** {avg_mem} MB  |  **Peak:** {peak_mem} MB")
    md.append(f"- **Avg CPU:** {avg_cpu}%  |  **Peak:** {peak_cpu}%\n")

    md.append("## 2. Contextual Action Spans (AI Breakpoint Analysis)")
    md.append("This table tracks execution latency *between* two marked events.\n")
    md.append("| Breakpoint A → B | Frames | Janky | Jank% | Health | Peak UI | Peak Raster | Status |")
    md.append("|---|---|---|---|---|---|---|---|")
    
    spans = calculate_action_spans()
    if not spans:
        md.append("| N/A | - | - | - | - | - | - | No events found |")
    else:
        for sp in spans:
            label = f"`{sp['start_event']}` → `{sp['end_event']}`"
            md.append(f"| {label} | {sp['duration_frames']} | {sp['janky_frames']} | {sp['jank_ratio']}% | {sp['span_health']}/100 | {sp['peak_ui']}ms | {sp['peak_raster']}ms | {sp['status']} |")
    md.append("\n")

    md.append("---\n## 🎯 Targeted AI Refactor Commands (Copy & Paste to Cursor)")
    md.append("The following commands have been dynamically generated based on the highest performance bottlenecks detected in your session.\n")
    
    bad_spans = [sp for sp in spans if "JANK" in sp["status"]]
    
    if not bad_spans:
        md.append("> ✅ **Status:** All tracked action spans completed within the 16.6ms budget. No targeted refactoring required.")
    else:
        bad_spans = sorted(bad_spans, key=lambda x: max(x['peak_ui'], x['peak_raster']), reverse=True)
        
        for i, sp in enumerate(bad_spans[:5]):
            md.append(f"### Priority Bottleneck #{i+1}")
            if sp['peak_ui'] > sp['peak_raster']:
                bottleneck_type = "UI Thread / Dart Logic (CPU)"
                diagnosis = "The UI thread choked. This means Dart code took too long to execute. I need you to check for deep/unnecessary widget rebuilds, heavy synchronous JSON parsing, large list iterations, or unoptimized `build()` methods."
            else:
                bottleneck_type = "Raster Thread / Skia (GPU)"
                diagnosis = "The Raster thread choked. The Dart logic was fine, but the GPU struggled to paint the frame. I need you to check for missing `RepaintBoundary` wrappers, expensive `SaveLayer` operations (like excessive `Opacity` or `ClipRRect`), or heavy image decoding."
            
            prompt = (
                f"> `@workspace` **Performance Regression Detected:** Between the exact moment `{sp['start_event']}` was fired and `{sp['end_event']}` was fired, the game suffered a **{bottleneck_type}** spike hitting **{max(sp['peak_ui'], sp['peak_raster'])}ms** (Budget is 16.6ms). \n>\n"
                f"> **Your Task:** Trace the execution path and Widget Tree transitions that bridge these two breakpoints. {diagnosis} Please rewrite or optimize the responsible Dart code to bring this span back under 16ms."
            )
            md.append(prompt)
            md.append("\n")

    md_string = "\n".join(md)
    buffer = io.BytesIO()
    buffer.write(md_string.encode('utf-8'))
    buffer.seek(0)

    return send_file(
        buffer,
        mimetype="text/markdown",
        as_attachment=True,
        download_name=f"HEXA_Perf_Audit_{time.strftime('%Y%m%d_%H%M%S')}.md"
    )

def find_free_port(preferred=None):
    """Return a usable TCP port.

    Tries the preferred port first (env PORT or the passed value); if it is
    taken or unspecified, asks the OS for any free port by binding to port 0.
    """
    candidates = []
    env_port = os.environ.get("PORT")
    if env_port and env_port.isdigit():
        candidates.append(int(env_port))
    if preferred:
        candidates.append(preferred)

    for port in candidates:
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            s.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
            try:
                s.bind(("127.0.0.1", port))
                return port
            except OSError:
                print(f"[PORT] {port} is busy, searching for a free port...")

    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        s.bind(("127.0.0.1", 0))
        return s.getsockname()[1]


if __name__ == '__main__':
    port = find_free_port(preferred=8000)
    print("=" * 60)
    print("  Performance Observability Lab")
    print(f"  Server running at: http://127.0.0.1:{port}")
    print("  Press CTRL+C to stop")
    print("=" * 60, flush=True)
    # use_reloader=False so the auto-selected port stays stable
    # (the reloader would re-run this file and could pick a different port).
    app.run(host="127.0.0.1", port=port, debug=True, use_reloader=False)